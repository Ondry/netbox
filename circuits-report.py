import io
import smtplib
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from django import forms

import openpyxl

from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import WHITE
from openpyxl.styles import NamedStyle
from django.contrib.contenttypes.models import ContentType

from circuits.models import CircuitTermination
from circuits.models.circuits import Circuit


from dcim.models import Device, Site, Location, Cable, Interface, CableTermination
from extras.models import CustomFieldChoiceSet
from netbox.configuration import FIELD_CHOICES

from tenancy.models import Contact,ContactAssignment
from extras.scripts import Script, TextVar, StringVar, ScriptVariable
from utilities.string import title

"""
font = Font(name='Calibri',
                  size=11,
                  bold=False,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
"""
green_color_code = "FF00B050"
green_color = openpyxl.styles.colors.Color(rgb=green_color_code)
green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green_color)
green_style = NamedStyle(name='green',fill=green_fill)

light_green_color_code = "FF90EE90"
light_green = openpyxl.styles.colors.Color(rgb=light_green_color_code)
light_green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=light_green)
light_green_style = NamedStyle(name='light_green',fill=light_green_fill)

red_color_code = "FFFFCCCB"
red_color = openpyxl.styles.colors.Color(rgb=red_color_code)
red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=red_color)
red_style = NamedStyle(name='red',fill=red_fill)

orange_color_code = openpyxl.styles.colors.Color(rgb="FFFF7415")
orange_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=orange_color_code)
orange_style = NamedStyle(name='orange',fill=orange_fill)

yellow_color = openpyxl.styles.colors.Color(rgb='FFFFFFC5')
yellow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=light_green)
yellow_style = NamedStyle(name='yellow',fill=yellow_fill)

blue_color_code = "FF00B0F0"
blue = openpyxl.styles.colors.Color(rgb=blue_color_code)
blue_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=blue)
blue_style = NamedStyle(name='blue',fill=blue_fill)

cyan_color_code = "FF0C0C0C"
cyan = openpyxl.styles.colors.Color(rgb=cyan_color_code)
cyan_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=cyan)
cyan_style = NamedStyle(name='cyan',fill=cyan_fill)

gray_color_code = "FFC0C0C0"
gray = openpyxl.styles.colors.Color(rgb=gray_color_code)
gray_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=gray)
gray_style = NamedStyle(name='gray',fill=gray_fill)

black_color_code = "FF000000"
black = openpyxl.styles.colors.Color(rgb=black_color_code)
font = Font(name='Calibri',color=WHITE)
black_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=black)
black_style = NamedStyle(name='black',fill=black_fill,font=font)


# Функция для отправки email с вложением
def send_message(send_to,
                 subject,
                 body,
                 attachment,
                 attachment_file_name,
                 attachment_file_type,
                 from_email="<your default sender email address>"):

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = send_to #COMMASPACE.join(send_to)
    msg['Subject'] = subject



    part = MIMEBase('application', attachment_file_type)
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={attachment_file_name}')
    msg.attach(part)

    # Подключение к SMTP серверу
    with smtplib.SMTP('localhost', 25) as server:
        server.send_message(msg)

class EmailVar(ScriptVariable):
    """
    Free-form text data. Renders as a <textarea>.
    """
    form_field = forms.EmailField

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.field_attrs['widget'] = forms.Textarea

class CircuitsReport(Script):
    class Meta:
        name = "Circuits Report"
        description = "Script for generating providers channels info"

    send_to = EmailVar(
        label="email",
        description="Email адрес для отправки отчета",
        required=True,
    )

    def fetch_providers(self):
        # Создаем книгу excel
        workbook = openpyxl.Workbook()
        workbook.add_named_style(green_style)
        workbook.add_named_style(light_green_style)
        workbook.add_named_style(yellow_style)
        workbook.add_named_style(red_style)
        workbook.add_named_style(blue_style)
        workbook.add_named_style(black_style)
        workbook.add_named_style(gray_style)
        workbook.add_named_style(cyan_style)
        workbook.add_named_style(orange_style)

        # Извлекаем лист из книги для работы с ним
        sheet = workbook.active

        # Записываем первой строкой названия столбцов
        fieldnames = [
            "ID",
            "Регион",
            "Учреждение(сайт)",
            "Статус учреждения",
            "Провайдер",
            "Договор(аккаунт)",
            "Приоритет канала",
            "Идентификатор канала связи",
            "Адрес предоставления услуги",
            "Статус канала",
            "Дата подключения",
            "Дата отключения",
            "Номер Д/С.",
            "Номер Б/З",
            "Дата Б/З (установки)",
            "Название роутера",
            "Название интерфейса роутера",
            "Скорость канала",
            "Скорость burst",
            "Лимит по трафику",
            "Фактическая скорость",
            "Статический белый адрес",
            "Технология подключения",
            "Стоимость подключения",
            "Ежемесячная плата",
            "Контакты",
            "Комментарии",
            "Описание",
            "Дата создания",
            "Дата последнего изменения",
        ]
        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 26
        sheet.column_dimensions["C"].width = 55
        sheet.column_dimensions["D"].width = 14
        sheet.column_dimensions["E"].width = 25
        sheet.column_dimensions["F"].width = 19
        sheet.column_dimensions["G"].width = 12
        sheet.column_dimensions["H"].width = 26
        sheet.column_dimensions["I"].width = 45
        sheet.column_dimensions["J"].width = 10
        sheet.column_dimensions["K"].width = 11
        sheet.column_dimensions["L"].width = 10
        sheet.column_dimensions["M"].width = 14
        sheet.column_dimensions["N"].width = 8
        sheet.column_dimensions["O"].width = 11
        sheet.column_dimensions["P"].width = 22
        sheet.column_dimensions["Q"].width = 8
        sheet.column_dimensions["R"].width = 8
        sheet.column_dimensions["S"].width = 8
        sheet.column_dimensions["T"].width = 8
        sheet.column_dimensions["U"].width = 8
        sheet.column_dimensions["V"].width = 8
        sheet.column_dimensions["W"].width = 16
        sheet.column_dimensions["X"].width = 8
        sheet.column_dimensions["Y"].width = 8
        sheet.column_dimensions["Z"].width = 28
        sheet.column_dimensions["AA"].width = 22
        sheet.column_dimensions["AB"].width = 30
        sheet.column_dimensions["AC"].width = 18
        sheet.column_dimensions["AD"].width = 18

        sheet.auto_filter.ref = sheet.dimensions

        sheet.append(fieldnames)

        index = 1
        # Итерируем все каналы связи
        for circuit in Circuit.objects.all():
            index += 1
            circuit_site = None
            circuit_site_status = None
            region = ""
            circuit_contacts = None
            circuit_device = None
            circuit_device_port = None
            cable_termination_a = None
            cable_termination_b = None
            # Если у канала связи есть точка подключения А, определяем сайт, регион, статус сайта,
            # а также роутер и порт к которому привязан канал связи
            if circuit.termination_a:
                # Определяем id привязанного сайта
                circuit_termination_a=circuit.termination_a
                circuit_termination_a_object_id = circuit.termination_a.termination_id
                # self.log_info(f"Fetching info for termination A id\n{circuit.termination_a}")
                circuit_termination_a_type = str(circuit.termination_a.termination_type)
                # Если тип соединенного объекта - сайт, определяем данные сайта привязанного к договору
                if circuit_termination_a_type == "DCIM | site":
                    try:
                        circuit_site = Site.objects.get(id=circuit_termination_a_object_id)
                        region = circuit_site.region.name
                        circuit_site_status = circuit_site.status
                        circuit_site = circuit_site.name
                    except Site.DoesNotExist:
                        self.log_failure(f"Сайт с id {circuit_termination_a_object_id} не найден.")
                # Если тип соединенного объекта - локация, определяем данные локации привязанного к договору
                elif circuit_termination_a_type == "DCIM | location":
                    try:
                        circuit_site_location = Location.objects.get(id=circuit_termination_a_object_id)
                        circuit_site_status = circuit_site_location.status
                        circuit_site_id = circuit_site_location.site.id
                        circuit_site = Site.objects.get(id=circuit_site_id)
                        region = circuit_site.region.name
                        circuit_site = f"{circuit_site.name} локация ({circuit_site_location.name})"
                    except Site.DoesNotExist:
                        self.log_failure(f"Локация сайта с id {circuit_termination_a_object_id} не найдена.")
                else:
                    self.log_failure(f"Неизвестный тип '{circuit_termination_a_type}' привязанного объекта к каналу связи")

                # Тут определяем роутер и порт, цепочка определния такая:
                # 1) По id канала(4315) ищем точку терминации А(2007) (выше уже она определена)
                # 2) По id точки треминации канала А ищем точку терминации кабеля
                # 3) по id кабеля из точки терминации кабеля А ищем точку терминации кабеля B
                # 4) Из точки терминации B получаем данные интерфейса, и из него же название роутера

                try:
                    self.log_info(f"Ищем точку подключения кабеля, к \n"
                                   f"circuit_termination_a = {circuit_termination_a}({circuit_termination_a.id}),\n"
                                   f"termination_type = {ContentType.objects.get_for_model(CircuitTermination)},\n"
                                   f"cable_end=A")
                    cable_termination_a = CableTermination.objects.get(termination_id=circuit_termination_a.id,
                                                                       termination_type = ContentType.objects.get_for_model(CircuitTermination),
                                                                       cable_end="A")
                    self.log_info(f"Найдена точка {cable_termination_a}")
                    self.log_info(f"Ищем точку подключения кабеля, к \n"
                                   f"circuit_termination_a_object_id = {cable_termination_a.cable},\n"
                                   f"termination_type = {ContentType.objects.get_for_model(Interface)},\n"
                                   f"cable_end=B")
                    # cable = Cable.objects.get(id=cable_termination_a.cable.id)
                    # cable_termination_b = cable.b_terminations[0]
                    cable_termination_b = CableTermination.objects.get(cable=cable_termination_a.cable,
                                                                       cable_end="B")
                    self.log_info(f"Найдена точка {cable_termination_b}")
                    circuit_device = cable_termination_b.termination.device.name
                    self.log_info(f"Определено устройство канада связи = {circuit_device}")
                    circuit_device_port = cable_termination_b.termination.name
                    self.log_info(f"Определен порт устройства канада связи = {circuit_device_port}")
                except CableTermination.DoesNotExist:
                    self.log_warning(f"Для точки терминации '{circuit.termination_a}({circuit_termination_a_object_id})' канала {circuit}({circuit.id}),"
                                     f" не удалось определить устройство и порт привязки,\n"
                                     f"cable_termination_a = {cable_termination_a},\n"
                                     f"cable_termination_b = {cable_termination_b},\n"
                                     f"circuit_device = {circuit_device},\n"
                                     f"circuit_device_port = {circuit_device_port}")
                    circuit_device = "Устройство не привязано"
                    circuit_device_port = "-"
                #except Exception as e:
                #     self.log_failure(f"Ошибка получения устройства и порта, через терминации кабеля, текст ошибки: {e}")
            else:
                self.log_warning(f"Канал связи {circuit}({circuit.id}) не привязан точкой А к сайту.")
                circuit_device = "Устройство не привязано"
                circuit_device_port = "-"

            try:
                contacts_ass_objects = (ContactAssignment.objects.filter(object_type=ContentType.objects.get_for_model(Circuit),
                                                                         object_id=circuit.id))
            except ContactAssignment.DoesNotExist:
                contacts_ass_objects = None

            circuit_contacts = ""
            if contacts_ass_objects:
                for contact_ass_object in contacts_ass_objects:
                    contact = Contact.objects.get(id=contact_ass_object.contact.id)
                    if contact:
                        str_contact = ""
                        str_contact += "---------------------\n"
                        if contact.name:
                            str_contact += f"Имя: {contact.name},\n"
                        if contact.group:
                            if contact.group.name:
                                str_contact += f"Группа контакта: {contact.group.name},\n"
                        if contact.phone:
                            str_contact += f"Телефон: {contact.phone},\n"
                        if contact.email:
                            str_contact += f"Email: {contact.email}\n"
                        if contact.description:
                            str_contact += f"Описание: {contact.description},\n"
                        circuit_contacts += str_contact

            install_date = circuit.install_date
            if install_date and type(install_date) == datetime:
                install_date = install_date.replace(tzinfo = None)

            termination_date = circuit.termination_date
            if termination_date and type(termination_date) == datetime:
                termination_date = termination_date.replace(tzinfo = None)

            created_date = circuit.created
            if created_date and type(created_date) == datetime:
                created_date = created_date.replace(tzinfo = None)

            updated_date = circuit.last_updated
            if updated_date and type(updated_date) == datetime:
                updated_date = updated_date.replace(tzinfo = None)

            """
                ID +
                Идентификатор канала связи +
                Провайдер +
                Аккаунт +
                - Тип
                Статус +
                - Арендатор
                - Сторона А
                - Сторона Z
                Описание +
                - Группа арендаторов
                Установлен +
                Разобран +
                Гарантированная скорость +
                Расширение по burst мбит / с
                Комментарии+
                Контакты -
                - Теги
                - Задания
                - Расстояние
                Адрес предоставления канала связи +
                Наименование учреждения +
                Фактическая скорость по замерам+
                Ограничение по графику в ГБ +
                - Идентификатор WAN	
                Стоимость подключения +	
                Ежемесячная стоимость +	
                nalog_nds
                Номер бланк заказа +	
                Номер Доп. Соглашения +	
                channel_ip_is_static +	
                Приоритет канала +	
                - Протокол подключения к оператору	
                Технология подключения +
                Создан(а) +
                Последнее обновление +
            """


            #"ID":
            sheet[f"A{index}"].value = str(circuit.id)

            #
            sheet[f"A{index}"].hyperlink = f"https://<your netbox url>/circuits/circuits/{circuit.id}/"

            # "Регион":
            sheet[f"B{index}"].value = region


            # "Учреждение(сайт)"
            sheet[f"C{index}"].value = circuit_site


            # "Статус учреждения"
            for choise in FIELD_CHOICES['dcim.Site.status']:
                if choise[0] == circuit_site_status:
                    sheet[f"D{index}"].value = choise[1]
                    sheet[f"D{index}"].style = choise[2]


            # "Провайдер"
            sheet[f"E{index}"].value = str(circuit.provider.name)


            # "Договор(аккаунт)"
            sheet[f"F{index}"].value = str(circuit.provider_account.account)


            # "Приоритет канала"
            priority = circuit.cf["channel_priority"]
            for choise in CustomFieldChoiceSet.objects.get(name="channel_priority").extra_choices:
                if choise[0] == priority:
                    sheet[f"G{index}"].value = choise[1]


            # "Идентификатор канала связи"
            sheet[f"H{index}"].value = str(circuit.cid)

            # "Адрес предоставления услуги"
            sheet[f"I{index}"].value = circuit.cf["channel_address"]

            sheet[f"I{index}"].alignment = Alignment(wrap_text=True)

            # "Статус канала"
            status = str(circuit.status)
            for choise in FIELD_CHOICES['circuits.Circuit.status']:
                if choise[0] == status:
                    sheet[f"J{index}"].value = choise[1]
                    sheet[f"J{index}"].style = choise[2]

            # "Дата подключения"
            sheet[f"K{index}"].value = install_date

            # "Дата отключения"
            sheet[f"L{index}"].value = termination_date

            # "Номер Д/С."
            sheet[f"M{index}"].value = circuit.cf["sub_account_name"]

            # "Номер Б/З"
            sheet[f"N{index}"].value = circuit.cf["order_name"]

            # "Дата Б/З (установки)"
            sheet[f"O{index}"].value = install_date


            # "Название роутера"
            if circuit_device == "Устройство не привязано":
                sheet[f"P{index}"].style = "red"
            else:
                sheet[f"P{index}"].style = "green"
            sheet[f"P{index}"].value = circuit_device

            # "Название интерфейса роутера"
            if circuit_device_port == "-" or circuit_device_port == "":
                sheet[f"Q{index}"].style = "red"
            else:
                sheet[f"Q{index}"].style = "green"
            sheet[f"Q{index}"].value = circuit_device_port

            # "Скорость канала"
            sheet[f"R{index}"].value = circuit.commit_rate

            # "Скорость burst"
            sheet[f"S{index}"].value = circuit.cf["burst"]

            # "Лимит по трафику"
            sheet[f"T{index}"].value = circuit.cf["speed_limit"]

            # "Фактическая скорость"
            sheet[f"U{index}"].value = circuit.cf["speed_fact"]

            # "Статический белый адрес"
            if circuit.cf["channel_ip_is_static"]:
                sheet[f"V{index}"].value = "Да"
                sheet[f"V{index}"].style = "green"

            # "Технология подключения"
            channel_type = circuit.cf["channel_type"]
            for choise in CustomFieldChoiceSet.objects.get(name="channel_type").extra_choices:
                if choise[0] == channel_type:
                    sheet[f"W{index}"].value = choise[1]

            # "Стоимость подключения"
            sheet[f"X{index}"].value = circuit.cf["install_price"]

            # "Ежемесячная плата"
            sheet[f"Y{index}"].value = circuit.cf["month_price"]

            # "Контакты"
            sheet[f"Z{index}"].value = circuit_contacts

            # "Комментарии"
            sheet[f"AA{index}"].value = str(circuit.comments)

            # "Описание"
            sheet[f"AB{index}"].value = str(circuit.description)
            sheet[f"AB{index}"].alignment = Alignment(wrap_text=True)

            # "Дата создания"
            sheet[f"AC{index}"].value = created_date

            # "Дата последнего изменения"
            sheet[f"AD{index}"].value = updated_date

        return workbook



    def run(self, data, commit):
        workbook = self.fetch_providers()

        now_time = datetime.now().strftime(format="%Y-%m-%d-%H-%M-%S")
        filename = f"{now_time}-provider_report.xlsx"

        # Create buffer in memory
        buffer = io.BytesIO()
        workbook.save(buffer)
        attachment = buffer.getvalue()

        send_message(send_to=data["send_to"],
                     from_email="<your sender email address>",
                     subject=f"Providers Report {now_time}",
                     body="",
                     attachment=attachment,
                     attachment_file_name=filename,
                     attachment_file_type="application/vnd.ms-excel")
        buffer.close()

        response = f"Report {filename} successfully generated and sent to recipients {data['send_to']}"

        return response
